import pandas as pd
import numpy as np
import time
import requests
from openpyxl.styles import Alignment, Font, PatternFill

# URL et paramètres
url = "https://archive-api.open-meteo.com/v1/archive"
params = {
    "latitude": 52.52,
    "longitude": 13.41,
    "start_date": "2020-01-01",
    "end_date": "2020-01-01",
    "hourly": "temperature_2m,relative_humidity_2m,dew_point_2m,precipitation,pressure_msl,surface_pressure,et0_fao_evapotranspiration,wind_speed_10m,wind_direction_10m,direct_radiation,direct_radiation_instant",
    "timezone" : "auto"
}



# Récupération des données
response = requests.get(url, params=params)
data = response.json()
print(data)
print(len(data["hourly"]["time"]))
print(data["hourly"]["temperature_2m"])
print(data["hourly"]["relative_humidity_2m"])

start_time = time.time()

colonnes_automate = ['NUMERO_E', 'HEURE_E', 'MINUTE_E', 'SECONDE_E', 'TEMP_CEL', 'REGUL_ECORIUM', 'T_ECOR_FOND', 'T_ECOR_INF', 'T_ECOR_SUP', 'REGUL_HYG', 'HYG_CEL', 'REGUL_PRESSION', 'PRESSION', 'REGUL_CO', 'CO', 'REGUL_O', 'O', 'REGUL_N', 'N', 'REGUL_X', 'X', 'HAUTEUR_ECOLUX', 'INTENSITE_ECOLUX', 'CONFIG_ECOLUX', 'PLUIE', 'VITESSE_VENT_PRINC', 'VENT_LATERAL', 'RENO_AIR', 'CONFIG_MES_INCUB', 'ACTION_UN', 'ACTION_DEUX', 'ACTION_TROIS', 'ACTION_QUATRE', 'ACTION_CINQ', 'ACTION_SIX', 'ACTION_SEPT', 'ACTION_HUIT', 'ACTION_NEUF', 'ACTION_DIX', 'ACTION_ONZE', 'ACTION_DOUZE', 'ACTION_TREIZE', 'ACTION_QUATORZE', 'ACTION_QUINZE', 'TEMP_PLUIE', 'CHROMATO', 'EV_VIDANGE_ECORIUM', 'EV_VIDANGE_CANIVEAU', 'EV_VIDANGE_CLIM', 'CONSIGNE_ESSAI', 'P_VAP_SAT', 'PT_ROSEE', 'HS_PROGR', 'T_BALLON_CHAUD', 'T_BALLON_FROID', 'BALLON_CHAUD_REEL', 'BALLON_FROID_REEL', 'Temp_progr', 'M_A']
colonnes_programmation = ['No_Pas', 'mn', 'secondes', 'no h', 'no jours', 'TEMP_CEL', 'REGUL_ECORIUM', 'T_ECOR_FOND', 'T_ECOR_INF', 'T_ECOR_SUP', 'REGUL_HYG', 'HYG_CEL', 'REGUL_PRESSION', 'PRESSION', 'REGUL_C', 'C', 'REGUL_O', 'O', 'REGUL_N', 'N', 'REGUL_X', 'X', 'HAUT_ECOLUX', 'INT_ECOLUX', 'CONFIG_ECOLUX', 'INT_PLUIE', 'VITESSE_VENT', 'VENT_LATERAL_UN', 'VENT_LATERAL_DEUX', 'VENT_LATERAL_TROIS', 'VENT_LATERAL_QUATRE', 'INT_VENT_LAT', 'RENO_AIR', 'EVENT Chromato', 'Vide_1', 'CONFIG_MES_INCUB', 'ACTION_1', 'ACTION_2', 'ACTION_3', 'ACTION_4', 'ACTION_5', 'ACTION_6', 'ACTION_7', 'ACTION_8', 'ACTION_9', 'ACTION_10', 'ACTION_11', 'ACTION_12', 'ACTION_13', 'ACTION_14', 'ACTION_15', 'TEMP_PLUIE', 'CHROMATO', 'EV_VIDANGE_ECORIUM', 'VIDANGE_CANIVEAU', 'EV_VIDANGE_CLIM', 'Vide_2', 'MAX T', 'MIN T', 'MAX HR', 'MIN HR', 'tests valeurs Temp', 'tests valeurs Hum', 'MARCHE', 'HYG_Corrigée pour Temp < 0', 'Min Temp', 'Max Temp', 'Temp Mode Mesure']


n = len(data["hourly"]["time"])  # Tu pourras changer ça par le nombre de ligne plus tard
df_programmation = pd.DataFrame(index=range(n), columns=colonnes_programmation)
df_automate = pd.DataFrame(columns=colonnes_automate)
df_infos = pd.DataFrame()
df_config = pd.DataFrame()



# --- 1. NO PAS (Optimisé) ---
# Le premier est 1, les autres sont =A{ligne_précédente}+1
df_programmation.at[0, 'No_Pas'] = 1
if n > 1:
    indices_prec = np.arange(2, n + 1).astype(str) # Lignes 2 à n
    df_programmation.loc[1:, 'No_Pas'] = np.char.add("=A", indices_prec) + "+1"

# --- 2. MN & NO H & NO JOURS (Optimisé) ---
# On crée les indices une seule fois pour tout le monde
idx_excel = np.arange(2, n + 2).astype(str)

df_programmation['mn'] = np.char.add("=MOD(INT(C", idx_excel) + "/60),60)"
df_programmation['no h'] = np.char.add("=INT(C", idx_excel) + "/3600)"
df_programmation['no jours'] = np.char.add("=INT(D", idx_excel) + "/24)+1"

# --- 3. SECONDES (Optimisé) ---
df_programmation.at[0, 'secondes'] = 0
if n > 1:
    indices_prec = np.arange(2, n + 1).astype(str)
    df_programmation.loc[1:, 'secondes'] = np.char.add("=C", indices_prec) + "+300"

# --- 4. T_ECOR (Correction) ---
formules_f = np.char.add("=$F", idx_excel)
# On crée une colonne verticale (n lignes, 1 colonne)
colonne_verticale = formules_f[:, np.newaxis]
# On la duplique sur 3 colonnes pour correspondre à cols_ecor
# np.tile(tableau, (répétition_ligne, répétition_colonne))
bloc_formules = np.tile(colonne_verticale, (1, 3))
cols_ecor = ['T_ECOR_FOND', 'T_ECOR_INF', 'T_ECOR_SUP']
df_programmation.loc[:, cols_ecor] = bloc_formules

# Optionnel mais plus robuste pour 50k lignes
cols_parametres = ['TEMP_CEL', 'REGUL_ECORIUM', 'REGUL_HYG', 'HYG_CEL', 'REGUL_PRESSION', 'PRESSION', 'REGUL_C', 'C', 'REGUL_O', 'O', 'REGUL_N', 'N', 'REGUL_X', 'X', 'HAUT_ECOLUX', 'INT_ECOLUX', 'CONFIG_ECOLUX', 'INT_PLUIE', 'VITESSE_VENT', 'VENT_LATERAL_UN', 'VENT_LATERAL_DEUX', 'VENT_LATERAL_TROIS', 'VENT_LATERAL_QUATRE', 'INT_VENT_LAT', 'RENO_AIR', 'EVENT Chromato', 'Vide_1', 'CONFIG_MES_INCUB', 'ACTION_1', 'ACTION_2', 'ACTION_3', 'ACTION_4', 'ACTION_5', 'ACTION_6', 'ACTION_7', 'ACTION_8', 'ACTION_9', 'ACTION_10', 'ACTION_11', 'ACTION_12', 'ACTION_13', 'ACTION_14', 'ACTION_15', 'TEMP_PLUIE', 'CHROMATO', 'EV_VIDANGE_ECORIUM', 'VIDANGE_CANIVEAU', 'EV_VIDANGE_CLIM', 'Vide_2']
ma_liste_complete = [22.0, 0, 1, 60, 0, 0, 1, 450, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 60, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 20, 0, 0, 0, 0, 0.0]
df_programmation[cols_parametres] = ma_liste_complete

# 3. ON ÉCRASE UNIQUEMENT TEMP_CEL AVEC LES DONNÉES DE L'API
# Comme n = len(data["hourly"]["temperature_2m"]), les tailles correspondent parfaitement
df_programmation['TEMP_CEL'] = data["hourly"]["temperature_2m"]
df_programmation['HYG_CEL'] = data["hourly"]["relative_humidity_2m"]

###### Automate ######

# 3. Formules pour NUMERO_E (=Programmation!A2, A3, etc.)
df_automate['NUMERO_E'] = np.char.add("=Programmation!A", idx_excel)

# HEURE_E : Pointe vers la colonne D (no h) de Programmation
df_automate['HEURE_E'] = np.char.add("=Programmation!D", idx_excel)

# MINUTE_E : Pointe vers la colonne B (mn) de Programmation
df_automate['MINUTE_E'] = np.char.add("=Programmation!B", idx_excel)

# SECONDE_E : Pointe vers la colonne C (secondes) de Programmation
df_automate['SECONDE_E'] = np.char.add("=Programmation!C", idx_excel)

df_automate['TEMP_CEL'] = np.char.add("=Programmation!F", idx_excel) + "*10"

df_automate['REGUL_ECORIUM'] = np.char.add("=Programmation!G", idx_excel)






### Config #######



# Utilisation de ExcelWriter
with pd.ExcelWriter('Road_to_Ecolab.xlsx', engine='openpyxl') as writer:
    df_automate.to_excel(writer, sheet_name='Automate', index=False)
    df_programmation.to_excel(writer, sheet_name='Programmation', index=False)
    df_infos.to_excel(writer, sheet_name='Infos', index=False)
    df_config.to_excel(writer, sheet_name='Config', index=False)

    # --- 1. ACCÈS ET CONTENU ---
    ws_config = writer.sheets['Config']

    # --- 2. STYLES ET FUSION ---
    style_jaune_flash = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    style_gras = Font(bold=True)
    style_centre = Alignment(horizontal="center", vertical="center")

    # --- 3. ÉCRITURE ET STYLE POUR A4/B4 et A5/B5 ---
    # On crée une liste des cellules à styliser pour gagner du temps
    cellules_a_formater = ['A4', 'B4', 'A5', 'B5']

    ws_config['A4'] = "Anticipation (min)"
    ws_config['A5'] = "Fortes Chaleurs"
    ws_config['B4'] = 60
    ws_config['B5'] = 0

    for ref in cellules_a_formater:
        ws_config[ref].font = style_gras
    ws_config[ref].alignment = style_centre

    # --- 4. TITRE FUSIONNÉ ---
    ws_config.merge_cells('A7:B7')
    cellule_titre = ws_config['A7']
    cellule_titre.value = "Configuration Régulation ECORIUM"
    cellule_titre.fill = style_jaune_flash
    cellule_titre.font = style_gras
    cellule_titre.alignment = style_centre

    # --- 5. RÈGLE GÉNÉRALE POUR "TOUT CE QUI SUIT" ---
    # Si tu veux que TOUTES les colonnes A et B soient en gras/centré par défaut
    # à partir de maintenant, on boucle sur les colonnes :
    # for col_lettre in ['A', 'B']:
    #     for cell in ws_config[col_lettre]:
    #         cell.font = style_gras
    #         cell.alignment = style_centre

    # --- 6. ÉLARGISSEMENT ---
    ws_config.column_dimensions['A'].width = 35  # Large pour les textes
    ws_config.column_dimensions['B'].width = 15  # Pour les valeurs 60, 0...

    # Si tu veux aussi élargir Programmation (60+ colonnes)
    ws_prog = writer.sheets['Programmation']
    for col in ws_prog.columns:
        column_letter = col[0].column_letter
        ws_prog.column_dimensions[column_letter].width = 18

    # Si tu veux aussi élargir Automate
    ws_auto = writer.sheets['Automate']
    for col in ws_auto.columns:
        column_letter = col[0].column_letter
        ws_auto.column_dimensions[column_letter].width = 18

end_time = time.time()
execution_time = end_time - start_time


print(f"Temps d'exécution total : {execution_time:.4f} secondes")




print("Done")